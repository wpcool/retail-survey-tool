"""
市场调研分析模块
整合自 BusinessAnalysis/market_search.py

功能：
1. 根据商品属性计算调整后售价
2. 计算价格指数（本店售价 / 竞争店售价）
3. 按门店+二级分类汇总分析
4. 按三级分类+采购员汇总分析（原采购分析）
5. 导出格式化Excel
"""

from sqlalchemy.orm import Session
from sqlalchemy import func
from models import SurveyRecord, Product, SurveyItem, SurveyTask
import pandas as pd
from typing import List, Dict, Optional, Tuple
from datetime import datetime
import json
import re

# 商品属性调价系数
ATTRIBUTE_PRICE_FACTORS = {
    '畅销商品': 0.95,
    '快消商品': 0.96,
    '正常商品': 0.97,
    '低周转商品': 0.98,
    '滞销商品': 0.99
}


def get_attribute_factor(product_attribute: Optional[str]) -> float:
    """
    获取商品属性的调价系数
    默认为正常商品 0.97
    """
    if not product_attribute:
        return 0.97
    return ATTRIBUTE_PRICE_FACTORS.get(product_attribute, 0.97)


def calculate_adjusted_price(sale_price: Optional[float], product_attribute: Optional[str]) -> Optional[float]:
    """
    根据商品属性计算调整后售价
    属性系数 * 本店参考售价
    
    Args:
        sale_price: 本店参考售价
        product_attribute: 商品属性（畅销/快消/正常/低周转/滞销）
    
    Returns:
        调整后售价
    """
    if not sale_price:
        return None
    
    factor = get_attribute_factor(product_attribute)
    return sale_price * factor


def calculate_price_index(our_price: Optional[float], competitor_price: Optional[float]) -> Optional[float]:
    """
    计算价格指数 = 本店售价 / 竞争店售价
    
    Args:
        our_price: 本店售价
        competitor_price: 竞争店售价
    
    Returns:
        价格指数，大于1表示本店更贵
    """
    if not our_price or not competitor_price or competitor_price == 0:
        return None
    return our_price / competitor_price


def parse_promotion_price(promotion_info: Optional[str]) -> Optional[float]:
    """
    从促销信息中提取促销价
    支持格式：直接数字（5.99）或包含数字的文本
    
    Args:
        promotion_info: 促销信息文本
    
    Returns:
        提取到的促销价，如果没有则返回None
    """
    if not promotion_info:
        return None
    
    # 尝试直接转换（如果是纯数字）
    try:
        return float(promotion_info)
    except ValueError:
        pass
    
    # 从文本中提取数字（如"促销价5.99元"）
    numbers = re.findall(r'\d+\.?\d*', str(promotion_info))
    if numbers:
        try:
            return float(numbers[0])
        except ValueError:
            pass
    
    return None


def fetch_survey_data(
    db: Session,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    own_store_name: Optional[str] = None,
    surveyor_id: Optional[int] = None
) -> pd.DataFrame:
    """
    从数据库获取调研数据并关联商品信息
    
    Returns:
        DataFrame包含以下列：
        - record_id, created_at, own_store_name, competitor_store_name
        - product_id, product_name, barcode
        - category_level1-4_name
        - brand_name, purchaser, product_attribute
        - purchase_price, sale_price, adjusted_price
        - competitor_price, promotion_price
        - price_index, promotion_price_index
    """
    # 构建基础查询
    query = db.query(
        SurveyRecord.id.label('record_id'),
        SurveyRecord.created_at,
        SurveyRecord.own_store_name,
        SurveyRecord.store_name.label('competitor_store_name'),
        SurveyRecord.price.label('competitor_price'),
        SurveyRecord.promotion_info,
        SurveyRecord.surveyor_id,
        SurveyItem.product_name.label('survey_product_name'),
        Product.id.label('product_id'),
        Product.barcode,
        Product.category_level1_name,
        Product.category_level2_name,
        Product.category_level3_name,
        Product.category_level4_name,
        Product.brand_name,
        Product.purchaser,
        Product.product_attribute,
        Product.purchase_price,
        Product.sale_price,
    ).join(
        SurveyItem, SurveyRecord.item_id == SurveyItem.id
    ).outerjoin(
        Product, SurveyItem.product_name == Product.name
    )
    
    # 应用筛选条件
    if start_date:
        query = query.filter(SurveyRecord.created_at >= f"{start_date} 00:00:00")
    if end_date:
        query = query.filter(SurveyRecord.created_at <= f"{end_date} 23:59:59")
    if own_store_name:
        query = query.filter(SurveyRecord.own_store_name == own_store_name)
    if surveyor_id:
        query = query.filter(SurveyRecord.surveyor_id == surveyor_id)
    
    # 执行查询
    results = query.all()
    
    # 调试统计
    missing_sale_price = 0
    missing_competitor_price = 0
    missing_product_match = 0
    total_records = len(results)
    
    # 转换为DataFrame
    data = []
    for r in results:
        # 检查商品关联是否成功
        if r.product_id is None:
            missing_product_match += 1
        
        # 检查本店售价是否存在
        if not r.sale_price:
            missing_sale_price += 1
        
        # 检查竞争店售价是否存在
        if not r.competitor_price:
            missing_competitor_price += 1
        
        # 计算调整后售价
        adjusted_price = calculate_adjusted_price(r.sale_price, r.product_attribute)
        
        # 解析促销价
        promotion_price = parse_promotion_price(r.promotion_info)
        
        # 计算调前价格指数（本店售价/竞争店售价）
        price_index = calculate_price_index(r.sale_price, r.competitor_price)
        
        # 计算调后价格指数（调整后售价/竞争店售价）
        adjusted_price_index = calculate_adjusted_price(r.sale_price, r.product_attribute)
        if adjusted_price_index:
            adjusted_price_index = calculate_price_index(adjusted_price_index, r.competitor_price)
        
        # 计算促销价价格指数
        promotion_price_index = calculate_price_index(r.sale_price, promotion_price)
        
        data.append({
            'record_id': r.record_id,
            'created_at': r.created_at,
            'date': r.created_at.strftime('%Y-%m-%d') if r.created_at else None,
            'own_store_name': r.own_store_name or '未指定',
            'competitor_store_name': r.competitor_store_name,
            'surveyor_id': r.surveyor_id,
            'product_id': r.product_id,
            'barcode': r.barcode,
            'product_name': r.survey_product_name,
            'category_level1_name': r.category_level1_name or '未分类',
            'category_level2_name': r.category_level2_name or '未分类',
            'category_level3_name': r.category_level3_name or '未分类',
            'category_level4_name': r.category_level4_name or '未分类',
            'brand_name': r.brand_name or '未知品牌',
            'purchaser': r.purchaser or '未指定',
            'product_attribute': r.product_attribute or '正常商品',
            'purchase_price': r.purchase_price or 0,
            'sale_price': r.sale_price or 0,
            'adjusted_price': adjusted_price or 0,
            'competitor_price': r.competitor_price or 0,
            'promotion_price': promotion_price,
            'promotion_info': r.promotion_info,  # 促销信息
            'price_index': price_index,  # 调前价格指数
            'adjusted_price_index': adjusted_price_index,  # 调后价格指数
            'promotion_price_index': promotion_price_index,
        })
    
    # 输出调试信息
    if total_records > 0:
        print(f"\n[市场分析调试信息]")
        print(f"总调研记录数: {total_records}")
        print(f"未关联到商品库: {missing_product_match} ({missing_product_match/total_records*100:.1f}%)")
        print(f"缺少本店售价(sale_price): {missing_sale_price} ({missing_sale_price/total_records*100:.1f}%)")
        print(f"缺少竞争店售价(price): {missing_competitor_price} ({missing_competitor_price/total_records*100:.1f}%)")
        print(f"价格指数计算成功: {total_records - missing_sale_price - missing_competitor_price} 条")
    
    return pd.DataFrame(data)


def analyze_by_store_and_category2(
    db: Session,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    own_store_name: Optional[str] = None
) -> pd.DataFrame:
    """
    门店维度分析（按自己门店 + 二级分类汇总）
    
    对应原系统的"区域分析"，维度改为：
    - 自己门店（如：蔡家坡店）
    - 二级分类名称（category_level2_name）
    
    输出字段：
    - 日期、自己门店、二级分类
    - 本店供价(合计)、本店售价(合计)、调整后售价(合计)
    - 竞争店售价(合计/平均)、价格指数(平均)
    - 调前毛利率、调后毛利率
    - 市调单品数、市调单品数占总数比
    """
    # 获取基础数据
    df = fetch_survey_data(db, start_date, end_date, own_store_name)
    
    if df.empty:
        return pd.DataFrame()
    
    # 按【自己门店+二级分类】分组汇总
    grouped = df.groupby(['own_store_name', 'category_level2_name']).agg({
        'purchase_price': 'sum',
        'sale_price': 'sum',
        'adjusted_price': 'sum',
        'competitor_price': ['sum', 'mean'],
        'price_index': 'mean',
        'adjusted_price_index': 'mean',
        'product_name': 'nunique',  # 市调单品数
    }).reset_index()
    
    # 展平多级列名
    grouped.columns = [
        '自己门店', '二级分类',
        '本店供价', '本店售价', '调整后售价',
        '竞争店售价(合计)', '竞争店售价(平均)', 
        '调前价格指数', '调后价格指数', '市调单品数'
    ]
    
    # 计算毛利率
    grouped['调前毛利率'] = grouped.apply(
        lambda row: (row['本店售价'] - row['本店供价']) / row['本店售价'] 
        if row['本店售价'] > 0 else 0, axis=1
    )
    grouped['调后毛利率'] = grouped.apply(
        lambda row: (row['调整后售价'] - row['本店供价']) / row['调整后售价'] 
        if row['调整后售价'] > 0 else 0, axis=1
    )
    
    # 计算市调单品数占总数比（需要获取该分类下的总商品数）
    # 简化处理：先计算当前占比，后续可以关联商品库完善
    total_products_by_cat2 = df.groupby('category_level2_name')['product_name'].nunique().to_dict()
    grouped['市调单品数占总数比'] = grouped.apply(
        lambda row: row['市调单品数'] / total_products_by_cat2.get(row['二级分类'], 1)
        if row['二级分类'] in total_products_by_cat2 else 0, axis=1
    )
    
    # 收集每个分组的详细商品列表（用于tooltip展示）
    product_details = {}
    for (store, cat2), group_df in df.groupby(['own_store_name', 'category_level2_name']):
        key = f"{store}|{cat2}"
        product_list = []
        for _, row in group_df.iterrows():
            product_list.append({
                'product_name': row['product_name'],
                'barcode': row['barcode'],
                'competitor_store': row['competitor_store_name'],
                'competitor_price': row['competitor_price'],
                'promotion_info': row['promotion_info'],
                'sale_price': row['sale_price'],
                'adjusted_price': row['adjusted_price'],
                'price_index': row['price_index'],
            })
        product_details[key] = product_list
    
    # 将商品详情添加到分组结果
    grouped['商品详情'] = grouped.apply(
        lambda row: product_details.get(f"{row['自己门店']}|{row['二级分类']}", []),
        axis=1
    )
    
    # 排序：先按门店，再按分类
    grouped = grouped.sort_values(['自己门店', '二级分类'])
    
    return grouped


def analyze_by_category3_and_purchaser(
    db: Session,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    own_store_name: Optional[str] = None
) -> pd.DataFrame:
    """
    采购维度分析（按三级分类 + 采购员汇总）
    
    对应原系统的"采购分析"，维度改为：
    - 三级分类名称（category_level3_name）
    - 采购员（purchaser）
    
    输出字段：
    - 日期、三级分类、采购员
    - 本店供价(合计)、本店售价(合计)、调整后售价(合计)
    - 竞争店售价(合计/平均)、价格指数(平均)
    - 调前毛利率、调后毛利率
    - 市调单品数、市调单品数占总数比
    """
    # 获取基础数据
    df = fetch_survey_data(db, start_date, end_date, own_store_name)
    
    if df.empty:
        return pd.DataFrame()
    
    # 按【三级分类+采购员】分组汇总
    grouped = df.groupby(['category_level3_name', 'purchaser']).agg({
        'purchase_price': 'sum',
        'sale_price': 'sum',
        'adjusted_price': 'sum',
        'competitor_price': ['sum', 'mean'],
        'price_index': 'mean',
        'adjusted_price_index': 'mean',
        'product_name': 'nunique',  # 市调单品数
    }).reset_index()
    
    # 展平多级列名
    grouped.columns = [
        '三级分类', '采购员',
        '本店供价', '本店售价', '调整后售价',
        '竞争店售价(合计)', '竞争店售价(平均)', 
        '调前价格指数', '调后价格指数', '市调单品数'
    ]
    
    # 计算毛利率
    grouped['调前毛利率'] = grouped.apply(
        lambda row: (row['本店售价'] - row['本店供价']) / row['本店售价'] 
        if row['本店售价'] > 0 else 0, axis=1
    )
    grouped['调后毛利率'] = grouped.apply(
        lambda row: (row['调整后售价'] - row['本店供价']) / row['调整后售价'] 
        if row['调整后售价'] > 0 else 0, axis=1
    )
    
    # 计算市调单品数占总数比
    total_products_by_cat3 = df.groupby('category_level3_name')['product_name'].nunique().to_dict()
    grouped['市调单品数占总数比'] = grouped.apply(
        lambda row: row['市调单品数'] / total_products_by_cat3.get(row['三级分类'], 1)
        if row['三级分类'] in total_products_by_cat3 else 0, axis=1
    )
    
    # 收集每个分组的详细商品列表（用于tooltip展示）
    product_details = {}
    for (cat3, purchaser), group_df in df.groupby(['category_level3_name', 'purchaser']):
        key = f"{cat3}|{purchaser}"
        product_list = []
        for _, row in group_df.iterrows():
            product_list.append({
                'product_name': row['product_name'],
                'barcode': row['barcode'],
                'competitor_store': row['competitor_store_name'],
                'competitor_price': row['competitor_price'],
                'promotion_info': row['promotion_info'],
                'sale_price': row['sale_price'],
                'adjusted_price': row['adjusted_price'],
                'price_index': row['price_index'],
            })
        product_details[key] = product_list
    
    # 将商品详情添加到分组结果
    grouped['商品详情'] = grouped.apply(
        lambda row: product_details.get(f"{row['三级分类']}|{row['采购员']}", []),
        axis=1
    )
    
    # 排序：先按分类，再按采购员
    grouped = grouped.sort_values(['三级分类', '采购员'])
    
    return grouped


def format_excel_output(file_path: str):
    """
    格式化Excel输出
    
    格式化规则：
    1. 毛利率、价格指数格式化为百分比
    2. 价格指数>0.95的单元格标红（警告）
    3. 自动调整列宽
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, numbers
    from openpyxl.utils import get_column_letter
    
    book = load_workbook(file_path)
    
    # 红色填充（用于价格指数>0.95的警告）
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    
    for sheet_name in book.sheetnames:
        ws = book[sheet_name]
        
        # 找到需要格式化的列
        percent_columns = ['调前毛利率', '调后毛利率', '调前价格指数', '调后价格指数', '市调单品数占总数比']
        price_index_cols = []
        
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header in ['调前价格指数', '调后价格指数']:
                price_index_cols.append(col_idx)
        
        # 格式化数据行
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                header = ws.cell(row=1, column=col_idx).value
                
                if header in percent_columns and cell.value is not None:
                    try:
                        value = float(cell.value)
                        cell.number_format = '0.00%'
                        cell.value = value
                        
                        # 价格指数>0.95标红
                        if header in ['调前价格指数', '调后价格指数'] and value > 0.95:
                            cell.fill = red_fill
                    except (ValueError, TypeError):
                        pass
                
                # 售价/供价列格式化为数字（不带¥符号）
                if header and (header in ['本店供价', '本店售价', '调整后售价', '竞争店售价(合计)', '竞争店售价(平均)']) and cell.value is not None:
                    try:
                        value = float(cell.value)
                        cell.number_format = '#,##0.00'
                        cell.value = value
                    except (ValueError, TypeError):
                        pass
        
        # 自动调整列宽
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    book.save(file_path)


def export_analysis_to_excel(
    df_store_analysis: pd.DataFrame,
    df_category_analysis: pd.DataFrame,
    output_path: str
):
    """
    导出分析结果到Excel
    
    包含两个Sheet：
    1. 门店分析（按自己门店+二级分类）
    2. 采购分析（按三级分类+采购员）
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 门店分析Sheet
        if not df_store_analysis.empty:
            df_store_analysis.to_excel(writer, sheet_name='门店分析', index=False)
        else:
            pd.DataFrame({'提示': ['暂无数据']}).to_excel(writer, sheet_name='门店分析', index=False)
        
        # 采购分析Sheet
        if not df_category_analysis.empty:
            df_category_analysis.to_excel(writer, sheet_name='采购分析', index=False)
        else:
            pd.DataFrame({'提示': ['暂无数据']}).to_excel(writer, sheet_name='采购分析', index=False)
    
    # 应用格式化
    format_excel_output(output_path)


def get_analysis_summary(
    db: Session,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    own_store_name: Optional[str] = None
) -> Dict:
    """
    获取分析汇总统计信息
    
    Returns:
        {
            'total_records': 总记录数,
            'total_stores': 涉及门店数,
            'total_products': 涉及商品数,
            'avg_price_index': 平均价格指数,
            'warning_count': 高价预警数（价格指数>0.95）
        }
    """
    df = fetch_survey_data(db, start_date, end_date, own_store_name)
    
    if df.empty:
        return {
            'total_records': 0,
            'total_stores': 0,
            'total_products': 0,
            'avg_price_index': 0,
            'warning_count': 0
        }
    
    return {
        'total_records': len(df),
        'total_stores': df['own_store_name'].nunique(),
        'total_products': df['product_name'].nunique(),
        'avg_price_index': df['price_index'].mean() if not df['price_index'].isna().all() else 0,
        'warning_count': len(df[df['price_index'] > 0.95])
    }
